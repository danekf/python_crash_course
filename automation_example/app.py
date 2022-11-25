#example that automates updating spreadsheets.

#import open excel workbook, and name it xl for easier reference
import openpyxl as xl

#add package to add bar charts
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
  #load the file
  wb = xl.load_workbook(filename)
  #load the sheet to work on
  sheet = wb['Sheet1']

  #loop through all rows in the sheet, starting at the one AFTER the titles
  for row in range(2, sheet.max_row + 1):
    #we want to affect the PRICE and update it, so select row3 and go to column 3 and set that as the target cell
    cell = sheet.cell(row, 3)
    print(cell.value)

    #adjust for 90% of price
    corrected_price = cell.value*0.9

    #add new column with new price rather than overwrite it in column 4
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price


  #add bar chart to workbook after the loop using the new data created in col 4 for all rows
  values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

  chart = BarChart()
  chart.add_data(values)

  #now that the chart is defined, add it using the created chart, at positiong e2
  sheet.add_chart(chart, 'e2')

  #save workbook
  wb.save(filename)


#and now call it
file = input('Enter Filename to update, including extension: ')
process_workbook(file)